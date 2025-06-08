#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
clean_business_data.py
~~~~~~~~~~~~~~~~~~~~~~
针对「司机端—控制面板_司机群账表2025.csv」做独立的数据清洗：
  1) 读取原始 CSV，做缺失值填充与格式统一（normalize_na、clean_contact、clean_dates、clean_area_other、clean_job_time、clean_vehicle、derive_tags）。
  2) 对核心业务字段做异常值标注（clean_business_core）。
  3) 输出清洗后结果到 CSV 和 Excel（含自动调整列宽与套用 Excel Table 样式）。
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional


import chardet
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import re
#客户评分列清洗函数，如果客户评分列存在非数字或空白内容，则替换为0
def clean_score_column(df: pd.DataFrame, col: str = "客户评分") -> pd.DataFrame:
    """
    把客户评分列所有非数字或空白内容都替换为0，支持小数。
    """
    if col in df.columns:
        df[col] = df[col].apply(lambda x: x if re.match(r"^\d+(\.\d+)?$", str(x)) else "0")
    return df

# ========== 配置 ==========
OUT_DIR = Path(r"E:\kabuda_data_analysis\司机业务信息库")
OUT_DIR.mkdir(parents=True, exist_ok=True)

RAW_FILE = Path(r"E:\kabuda_data_analysis\Lark\司机端—控制面板_司机群账表2025.csv")
CSV_OUT  = OUT_DIR / "司机业务.csv"
XLSX_OUT = OUT_DIR / "司机业务.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)

# ========== 1. 读取原始档案 ==========
def detect_encoding(fp: Path) -> str:
    with open(fp, "rb") as f:
        return chardet.detect(f.read())["encoding"] or "utf-8"

def load_raw(fp: Path) -> pd.DataFrame:
    ext = fp.suffix.lower()
    logging.info("加载原始档案: %s", fp)
    if ext == ".csv":
        enc = detect_encoding(fp)
        logging.info("检测到编码: %s", enc)
        return pd.read_csv(fp, encoding=enc, dtype=str)
    elif ext in (".xls", ".xlsx"):
        return pd.read_excel(fp, dtype=str, engine="openpyxl")
    else:
        raise ValueError(f"不支持的文件类型: {ext}")

# ========== 2. 基础清洗管道（尽量保证所有字段都有基本格式） ==========
def normalize_na(df: pd.DataFrame) -> pd.DataFrame:
    # 将空字符串、"nan"、"None" 等统一为 "缺失"
    for col in df.columns:
        df[col] = (
            df[col]
            .astype(str)
            .replace(["nan", "None", "NaN", ""], "缺失")
            .str.strip()
            .fillna("缺失")
        )
    return df

def clean_contact(df: pd.DataFrame) -> pd.DataFrame:
    # 如果表里没有手机号/邮箱/微信号等列，则跳过
    if "手机号" in df.columns:
        phone_re = re.compile(r"^1\d{10}$|^6\d{9}$")
        df["手机号"] = (
            df["手机号"].astype(str)
            .apply(lambda x: re.sub(r"\D", "", x))
            .apply(lambda x: x if phone_re.match(x) else "缺失")
        )
    if "邮箱" in df.columns:
        df["邮箱"] = df["邮箱"].apply(
            lambda x: x if re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", str(x)) else "缺失"
        )
    if "微信号" in df.columns:
        df["微信号"] = df["微信号"].apply(
            lambda x: x if 3 < len(str(x)) < 50 and x != "缺失" else "缺失"
        )
    return df

def clean_dates(df: pd.DataFrame) -> pd.DataFrame:
    # 对所有看起来像“时间”的列尝试转换
    for col in ["订单时间", "实际操作时间"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df

def clean_area_other(df: pd.DataFrame) -> pd.DataFrame:
    if "活动地区" in df.columns:
        df["活动地区"] = (
            df["活动地区"]
            .astype(str)
            .str.replace(r"其他\(请说明地区\)", "", regex=True)
            .str.strip()
            .replace({"": "缺失"})
        )
    return df

def clean_job_time(df: pd.DataFrame) -> pd.DataFrame:
    if "接单时间" in df.columns:
        df["接单时间"] = (
            df["接单时间"].astype(str)
            .str.replace(r"[，、]", ",", regex=True)
            .apply(lambda x: "缺失" if x == "缺失" else x.split(",")[0])
        )
    return df

def clean_vehicle(df: pd.DataFrame) -> pd.DataFrame:
    # 仅处理“车辆信息”“驾照”列，若存在
    if "车辆信息" in df.columns:
        df["车辆信息"] = df["车辆信息"].astype(str).apply(
            lambda x: "缺失" if x.strip() == "" else x
        )
    if "驾照" in df.columns:
        df["驾照"] = df["驾照"].astype(str).apply(
            lambda x: "缺失" if x.strip() == "" else x
        )
    return df

def derive_tags(df: pd.DataFrame) -> pd.DataFrame:
    # 示例：生成“是否含多地区”标签
    if "活动地区" in df.columns:
        df["含多地区"] = df["活动地区"].astype(str).str.contains(",").map({True: "是", False: "否"})
    return df
def clean_late_column(df: pd.DataFrame, col: str = "是否迟到") -> pd.DataFrame:
    """
    将“是否迟到”列：0→迟到，1→没有迟到，其它保持原样
    """
    if col in df.columns:
        df[col] = df[col].map({"0": "迟到", "1": "准时"}).fillna(df[col])
    return df

def run_basic_pipeline(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw.copy()
    df = normalize_na(df)
    df = clean_score_column(df)
    df = clean_late_column(df)

    df = clean_contact(df)
    df = clean_dates(df)
    df = clean_area_other(df)
    df = clean_job_time(df)
    df = clean_vehicle(df)
    df = derive_tags(df)
    return df

# ========== 3. 核心业务表异常值清洗函数 ==========
def clean_business_core(df: pd.DataFrame) -> pd.DataFrame:
    """
    对业务核心表做异常值处理，返回清洗后的 DataFrame。
    包含：
      - 日期字段校验与异常标注
      - 数值字段提取与校验
      - 分类字段校验
      - 文本/标识字段异常标注
    """

    df = df.copy()

    # --- 1. 日期字段处理 --- #
    if "订单时间" in df.columns:
        #     df["订单时间_原"] = df["订单时间"]
        #     df["订单时间"] = pd.to_datetime(df["订单时间"], errors="coerce")
        #     df.loc[df["订单时间"].isna(), "订单时间_异常"] = "无法解析"
        # if "实际操作时间" in df.columns:
        #     df["实际操作时间_原"] = df["实际操作时间"]
        #     df["实际操作时间"] = pd.to_datetime(df["实际操作时间"], errors="coerce")
        #     df.loc[df["实际操作时间"].isna(), "实际操作时间_异常"] = "无法解析"
        # # 时间顺序逻辑检查
        # if all(col in df.columns for col in ("订单时间", "实际操作时间")):
        #     mask_time = (
        #         df["订单时间"].notna()
        #         & df["实际操作时间"].notna()
        #         & (df["实际操作时间"] < df["订单时间"])
        #     )
        #     df.loc[mask_time, "时间逻辑异常"] = "实际操作早于下单"
        pass

    # --- 2. 数值字段处理 --- #
    def to_float(col: str) -> pd.Series:
        if col in df.columns:
            return pd.to_numeric(df[col].astype(str).str.replace(r"[^\d\.]", "", regex=True), errors="coerce")
        return pd.Series(dtype="float64")

    if "客户评分" in df.columns:
        # 1. 用已有的 to_float() 函数把字符串临时转成数值 Series
        score = to_float("客户评分")

        # 2. 定义异常掩码——非空且 <0 或 >10
        mask = score.notna() & ((score < 0) | (score > 10))

        # 3. 只在原列 “客户评分” 的这些位置写入 “缺失”
        df.loc[mask, "客户评分"] = "缺失"

        # （可选）把本来就是空或 NaN 的也标成“缺失”
        df.loc[df["客户评分"].isna() | (df["客户评分"] == ""), "客户评分"] = "缺失"


    df["订单销售价"] = to_float("订单销售价")
    df["司机应收"] = to_float("司机应收")
    df["公司应收"] = to_float("公司应收")
    # df["发票金额"] = to_float("发票金额")

    mask_mismatch = (
        df["订单销售价"].notna()
        & df["司机应收"].notna()
        & df["公司应收"].notna()
        & (df["订单销售价"] != df["司机应收"] + df["公司应收"])
    )
    df.loc[mask_mismatch, "金额不匹配"] = "司机应收+公司应收 ≠ 订单销售价"

    # mask_invoice = (
    #     df["发票金额_数值"].notna()
    #     & df["订单销售价_数值"].notna()
    #     & (df["发票金额_数值"] != df["订单销售价_数值"])
    # )
    # df.loc[mask_invoice, "发票金额_异常"] = "发票金额与订单价格不符"

    # --- 3. 分类字段校验 --- #
# —— 1. 标准化原始订单类型 ——
    # 原始可能的选项
    raw_types = [
        '宠管','代驾','接机/送机','接送',
        '跑腿','闪送','小程序接送机','其他'
    ]
    # 用“缺失”填空，并把非列表内值也当成“其他”
    df['订单类型'] = df['订单类型'].fillna('缺失').replace({'': '缺失'})
    df.loc[~df['订单类型'].isin(raw_types), '订单类型'] = '其他'

    # —— 2. 定义归类映射 ——
    mapping = {
        '接机/送机': '接送服务',
        '接送':     '接送服务',
        '小程序接送机': '接送服务',
        '跑腿':     '跑腿服务',
        '闪送':     '跑腿服务',
        '代驾':     '代驾服务',
        '宠管':     '宠物服务',
        '包车':     '包车服务',
        '其他':     '其他',
        '缺失':     '缺失'
    }
    df['订单类型'] = df['订单类型'].map(mapping)




    if "支付方式" in df.columns:
        valid_pay = {"微信", "支付宝", "现金"}
        df["支付方式"] = df["支付方式"].fillna("缺失").replace({"": "缺失"})
        df.loc[~df["支付方式"].isin(valid_pay), "支付方式"] = "其他"

    if "支付状态" in df.columns:
        valid_status = {"已支付", "未支付", "待支付"}
        df["支付状态"] = df["支付状态"].fillna("缺失").replace({"": "缺失"})
        df.loc[~df["支付状态"].isin(valid_status), "支付状态"] = "缺失"

    if "发票状态" in df.columns:
        valid_inv = {"已开票", "未开票"}
        df["发票状态"] = df["发票状态"].fillna("缺失").replace({"": "缺失"})
        df.loc[~df["发票状态"].isin(valid_inv), "发票状态"] = "缺失"

    if "订单状态" in df.columns:
        valid_ord = {"已完成", "进行中", "已取消", "已过期"}
        df["订单状态"] = df["订单状态"].fillna("缺失").replace({"": "缺失"})
        df.loc[~df["订单状态"].isin(valid_ord), "订单状态"] = "缺失"

    if "是否已结算" in df.columns:
        df["是否已结算"] = df["是否已结算"].fillna("缺失").replace({"": "缺失"})
        df.loc[~df["是否已结算"].isin({"是", "否"}), "是否已结算"] = "缺失"

    # --- 4. 文本/标识字段校验 --- #
    # ———— 订单编号处理 ————
    if "订单编号" in df.columns:
        # 标记缺失：NaN 或 空字符串
        missing_order = df["订单编号"].isna() | (df["订单编号"] == "")
        # 标记重复：所有重复且非缺失的行
        dup_mask = df["订单编号"].duplicated(keep=False) & ~missing_order

        # 先把重复的订单号替换为“订单编号重复”
        df.loc[dup_mask, "订单编号"] = "订单编号重复"
        # 再把缺失的订单号替换为“缺失”
        df.loc[missing_order, "订单编号"] = "缺失"


    # ———— 客户微信号处理 ————
    if "客户微信号" in df.columns:
        mask = (
            df["客户微信号"].isna()  # NaN
            | (df["客户微信号"] == "")  # 空字符串
            | df["客户微信号"]
                .str.contains(r'[\u4e00-\u9fa5]', regex=True, na=False)  # 包含汉字
        )
        df.loc[mask, "客户微信号"] = "缺失"


    # ———— 实际收款方处理 ————
    if "实际收款方" in df.columns:
        missing_payee = df["实际收款方"].isna() | (df["实际收款方"] == "")
        df.loc[missing_payee, "实际收款方"] = "缺失"


    # ———— 车辆信息处理 ————
    if "车辆信息" in df.columns:
        # strip() 去掉空白后判断是否为空
        missing_vehicle = df["车辆信息"].isna() | (df["车辆信息"].str.strip() == "")
        df.loc[missing_vehicle, "车辆信息"] = "缺失"


    if "收付款证明" in df.columns:
        pass # 这里可以添加收付款证明的处理逻辑

    df=process_orders(df)

    return df

# ========== 4. 自适应列宽 ==========
def auto_adjust_column_width(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        # 获取所有列的表头（第一行单元格值）
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for idx, col in enumerate(ws.columns):
            # 取每一列所有内容的最大长度，并加上列名长度
            values = [str(cell.value or "") for cell in col]
            header_len = len(str(headers[idx])) if idx < len(headers) else 0
            max_length = max([len(v) for v in values] + [header_len])
            col_letter = col[0].column_letter
            # print(f"列 {col_letter} 最大长度: {max_length}, 列名长度: {header_len}")
            ws.column_dimensions[col_letter].width = max_length + 4
    wb.save(xlsx_path)






# ========== 5. 套用 Excel Table 样式（可选） ==========
def apply_excel_table_styles(xlsx_path: Path) -> None:
    """
    将每个 sheet 的数据区域转换为 Excel Table，并应用 TableStyleMedium9 样式。
    """
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tbl = Table(displayName=f"{ws.title}_tbl", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
    wb.save(xlsx_path)

# ========== 6. 保存清洗结果 ==========
def save_cleaned(df: pd.DataFrame) -> None:
    # 写出 CSV
    logging.info("写出 CSV -> %s", CSV_OUT)
    df.to_csv(CSV_OUT, index=False, encoding="utf-8-sig")

    # 写出 Excel，仅一张 sheet “清洗后业务表”
    logging.info("写出 Excel -> %s", XLSX_OUT)
    with pd.ExcelWriter(XLSX_OUT, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name="清洗后业务表", index=False)



    # 自适应列宽
    logging.info("调整列宽自适应内容宽度 -> %s", XLSX_OUT)
    auto_adjust_column_width(XLSX_OUT)

    # 套用 Excel Table 样式（可选，如果不需要可注释掉）
    logging.info("为 sheet 应用 Excel Table 样式 -> %s", XLSX_OUT)
    apply_excel_table_styles(XLSX_OUT)



def print_col_width_info(df, colname):
    """
    打印某列的列名长度和内容最大长度
    """
    header_len = len(colname)
    content_max = df[colname].astype(str).map(len).max()
    print(f"列名长度: {header_len}")
    print(f"内容最大长度: {content_max}")


def process_orders(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理订单 DataFrame：
    1) 隐藏（删除）确认报单列
    2) 已确认保单列：1->"是"，0->"否"
    """
    # 1. 隐藏“确认报单”列
    if "确认报单" in df.columns:
        df = df.drop(columns="确认报单")

    # 2. 转换“已确认保单”列的数值
    # 1. 尝试把“已确认报单”转成数值，非 1/0 的都会被转成 NaN
    numeric = pd.to_numeric(df["已确认报单"], errors="coerce")

    # 2. 对转换失败（NaN）的行写入“缺失”
    df.loc[numeric.isna(), "已确认报单"] = "缺失"

    # 3. 对数值 1/0 再分别映射
    df.loc[numeric == 1, "已确认报单"] = "是"
    df.loc[numeric == 0, "已确认报单"] = "否"


    return df

# ========== 7. 主函数 ==========
def main():
    raw_df = load_raw(RAW_FILE)

    # 基础清洗管道
    basic_cleaned = run_basic_pipeline(raw_df)

    # 核心业务表异常值清洗
    business_cleaned = clean_business_core(basic_cleaned)

    # 保存结果
    save_cleaned(business_cleaned)



    logging.info(
        "清洗完成：共 %d 行记录。",
        len(business_cleaned)
    )

if __name__ == "__main__":
    main()

