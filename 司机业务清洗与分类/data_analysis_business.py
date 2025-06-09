#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
driver_business_quickstats_mix_v2_optimized.py
---------------------------------
区域/流向归一正则+LLM+缓存，详细日志。
"""

import os
import re
import sys
import json
from pathlib import Path
from dotenv import load_dotenv
import pandas as pd
import openai
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import logging

# ===== 日志配置 =====
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[logging.StreamHandler(sys.stderr)]
)

# ===== 环境与Key =====
load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")
MODEL = "gpt-3.5-turbo"

# ===== 路径配置 =====
BASE_DIR = Path(r"E:\kabuda_data_analysis\司机业务信息库")
CSV = BASE_DIR / "司机业务.csv"
XLSX = BASE_DIR / "司机业务.xlsx"
XLSX_OUT = BASE_DIR / "司机业务_统计分析.xlsx"
CACHE_PATH = BASE_DIR / "area_normalize_cache.json"

# ===== 数据读取 =====
if CSV.exists():
    df = pd.read_csv(CSV, dtype=str)
elif XLSX.exists():
    df = pd.read_excel(XLSX, sheet_name=0, dtype=str)
else:
    raise FileNotFoundError("未找到司机业务.csv或司机业务.xlsx")
df = df.fillna("")
logging.info(f"读取数据，行数: {len(df)}")

# ===== 区域归一：正则/手工映射 + LLM + 缓存 =====
AREA_MAP = {
    "Toronto": "多伦多", "多伦多": "多伦多", "toronto": "多伦多",
    "North York": "北约克", "北约克": "北约克", "northyork": "北约克",
    "Scarborough": "士嘉堡", "士嘉堡": "士嘉堡", "scarborough": "士嘉堡",
    "Markham": "万锦", "万锦": "万锦", "markham": "万锦",
    "Richmond Hill": "列治文山", "列治文山": "列治文山",
    "Mississauga": "密西沙加", "密西沙加": "密西沙加",
    "Etobicoke": "怡陶碧谷", "皮尔逊": "皮尔逊机场", "Pearson": "皮尔逊机场", "机场": "皮尔逊机场"
}
if CACHE_PATH.exists():
    with open(CACHE_PATH, "r", encoding="utf-8") as f:
        area_cache = json.load(f)
else:
    area_cache = {}

def normalize_area(text, biz_type="未知", log_prefix=""):
    """标准化单个地点字符串，自动日志输出"""
    orig_text = text.strip()
    # 先查映射
    for k, v in AREA_MAP.items():
        if k.lower() in orig_text.lower():
            norm = v
            break
    else:
        # 查缓存
        if orig_text in area_cache:
            norm = area_cache[orig_text]
        else:
            # LLM归一
            try:
                prompt = f"请将下述地址归一化为GTA地区常见的行政区名（如多伦多、北约克、士嘉堡、万锦、列治文山、密西沙加、皮尔逊机场等），仅返回地名，不加其它：\n{orig_text}"
                resp = openai.chat.completions.create(
                    model=MODEL,
                    messages=[{"role": "system", "content": prompt}],
                    temperature=0, max_tokens=8)
                norm = resp.choices[0].message.content.strip().replace("。", "")
                norm = norm if norm else orig_text
                area_cache[orig_text] = norm
                with open(CACHE_PATH, "w", encoding="utf-8") as f:
                    json.dump(area_cache, f, ensure_ascii=False, indent=2)
                logging.info(f"{log_prefix}LLM归一: 原始:「{orig_text}」→ 归一:「{norm}」| 业务类型: {biz_type}")
            except Exception as e:
                logging.error(f"{log_prefix}LLM归一失败: {e}")
                norm = orig_text
    if not log_prefix.startswith("[Flow]"):  # 避免流向日志太多
        logging.info(f"{log_prefix}区域归一: 原始:「{orig_text}」→ 归一:「{norm}」| 业务类型: {biz_type}")
    return norm

# ====== 结构化“订单概述”主要字段（原正则+LLM补齐）======
def extract_info(text):
    # 业务类型
    if re.search(r"接机|送机", text):
        type_ = "接送机"
    elif re.search(r"包车|一日游|多日游|包日", text):
        type_ = "包车"
    elif re.search(r"跑腿|代买|代取|代送|代购", text):
        type_ = "跑腿"
    elif re.search(r"行李寄存|寄存", text):
        type_ = "行李寄存"
    elif re.search(r"搬家|搬运", text):
        type_ = "搬家"
    elif re.search(r"电话|叫醒|叫人", text):
        type_ = "代办/其它"
    else:
        type_ = ""
    # 区域判定（先粗取）
    area_match = re.search(
        r"多伦多|Toronto|皮尔逊|Markham|Richmond Hill|万锦|Scarborough|士嘉堡|约克|North York|Etobicoke|密西沙加|Mississauga|机场",
        text, re.I)
    area_ = area_match.group(0) if area_match else ""
    # 金额
    amount_match = re.search(r"[💰\$](\d+(\.\d+)?)", text)
    amount = amount_match.group(1) if amount_match else ""
    # 起点终点
    addresses = re.findall(
        r"从\s*([\u4e00-\u9fa5a-zA-Z0-9 ,#\-]+?)(?:到|—|-|－|——)\s*([\u4e00-\u9fa5a-zA-Z0-9 ,#\-]+)",
        text)
    if addresses:
        start, end = addresses[0]
    else:
        start, end = "", ""
    # 时间
    time_match = re.search(
        r"(\d{1,2}[:：]\d{2}\s*(?:AM|PM|am|pm)?)|(\d{1,2}点半?)|(\d{1,2}/\d{1,2}\s*\d{1,2}[:：]\d{2})|(?:上午|下午|中午)\s*\d{1,2}[:：]?\d{0,2}",
        text)
    time_ = time_match.group(0) if time_match else ""
    return {
        "业务类型_struct": type_,
        "区域_struct": area_,
        "金额_struct": amount,
        "起点": start,
        "终点": end,
        "时间_struct": time_
    }

extract_results = df["订单概述"].apply(extract_info).apply(pd.Series)
to_llm_idx = extract_results[(extract_results["业务类型_struct"] == "") | (extract_results["区域_struct"] == "")].index
llm_targets = df.loc[to_llm_idx, "订单概述"].tolist()

def llm_extract(batch):
    out = []
    for text in batch:
        logging.info(f"[LLM-Extract] 原文: {text}")
        prompt = (
            "你是业务归类助手，请仅输出如下格式：\n"
            "业务类型: <类型>\n区域: <区域>\n"
            "只允许返回两行，不加其它文字。"
        )
        try:
            resp = openai.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": text},
                ],
                temperature=0.0, max_tokens=30,
            )
            ans = resp.choices[0].message.content.strip()
            t = re.search(r"业务类型[:：]\s*(\S+)", ans)
            a = re.search(r"区域[:：]\s*(\S+)", ans)
            biz = t.group(1) if t else ""
            area = a.group(1) if a else ""
            logging.info(f"[LLM-Extract] 归一结果: 业务类型: {biz}, 区域: {area}")
            out.append({
                "业务类型_struct": biz,
                "区域_struct": area
            })
        except Exception as e:
            logging.error(f"[LLM-Extract] call failed: {e}")
            out.append({"业务类型_struct": "", "区域_struct": ""})
    return pd.DataFrame(out)

if llm_targets:
    batches = [llm_targets[i:i+10] for i in range(0, len(llm_targets), 10)]
    llm_out = pd.concat([llm_extract(b) for b in batches], ignore_index=True)
    extract_results.loc[to_llm_idx, ["业务类型_struct", "区域_struct"]] = llm_out.values

# 合并回df
for col in extract_results.columns:
    df[col] = extract_results[col]

# ===== 区域归一：应用于所有相关字段 =====
df["区域归一"] = [
    normalize_area(area, biz_type=biz, log_prefix="[区域分布] ")
    for area, biz in zip(df["区域_struct"], df["业务类型_struct"])
]

# ====== 起点/终点归一（流向用）======
df["起点归一"] = [
    normalize_area(addr, biz_type=biz, log_prefix="[Flow-起点] ")
    for addr, biz in zip(df["起点"], df["业务类型_struct"])
]
df["终点归一"] = [
    normalize_area(addr, biz_type=biz, log_prefix="[Flow-终点] ")
    for addr, biz in zip(df["终点"], df["业务类型_struct"])
]

# ========== 字段融合 ==========

df["业务类型_结构化"] = df["业务类型_struct"]
if "订单类型" in df.columns:
    df.loc[df["业务类型_结构化"] == "", "业务类型_结构化"] = df["订单类型"]
df["业务类型_大类"] = df.get("订单类型", "")

# ========== 统计分析Sheet生成 ==========

def try_to_float(x):
    try:
        return float(x)
    except:
        return None

report_tables = {}

# 1. 业务类型分布_细分
vc = df["业务类型_结构化"].value_counts(dropna=False).reset_index()
vc.columns = ["业务类型_细分", "数量"]
report_tables["业务类型分布_细分"] = vc

# 2. 业务类型分布
vc = df["业务类型_大类"].value_counts(dropna=False).reset_index()
vc.columns = ["业务类型", "数量"]
report_tables["业务类型分布"] = vc

# 3. 区域分布（归一化后）
vc = df["区域归一"].value_counts(dropna=False).reset_index()
vc.columns = ["区域", "数量"]
report_tables["区域分布"] = vc

# 4. 金额区间分布
for col in ["金额_struct", "订单金额", "金额"]:
    if col in df.columns:
        amount = df[col].apply(try_to_float)
        bins = [0, 50, 100, 200, 500, 1000, float('inf')]
        labels = ["0-50", "50-100", "100-200", "200-500", "500-1000", "1000+"]
        cut = pd.cut(amount, bins=bins, labels=labels, right=False)
        vc = cut.value_counts(sort=False).reset_index()
        vc.columns = ["区间", "数量"]
        report_tables["金额区间分布"] = vc
        break

# 5. 订单状态分布
for col in ["订单状态"]:
    if col in df.columns:
        vc = df[col].value_counts(dropna=False).reset_index()
        vc.columns = [col, "数量"]
        report_tables["订单状态分布"] = vc
        break

# 6. 评分区间分布
for col in ["评分", "客户评分"]:
    if col in df.columns:
        score = df[col].apply(try_to_float)
        bins = [0, 3, 5, 8, 10]
        labels = ["0-3", "3-5", "5-8", "8-10"]
        cut = pd.cut(score, bins=bins, labels=labels, right=False, include_lowest=True)
        vc = cut.value_counts(sort=False).reset_index()
        vc.columns = ["区间", "数量"]
        report_tables["评分区间分布"] = vc
        break

# 7. 迟到分布
for col in ["是否迟到", "迟到", "迟到次数"]:
    if col in df.columns:
        vc = df[col].value_counts(dropna=False).reset_index()
        vc.columns = [col, "数量"]
        report_tables["迟到分布"] = vc
        break

# 8. 按月趋势
for col in ["下单时间", "订单日期", "创建时间"]:
    if col in df.columns:
        dates = pd.to_datetime(df[col], errors="coerce")
        monthly = dates.dt.to_period("M").value_counts().sort_index().reset_index()
        monthly.columns = ["月份", "订单数"]
        report_tables["每月订单趋势"] = monthly
        break

# 9. 司机分布
for col in ["司机", "司机姓名", "司机ID"]:
    if col in df.columns:
        vc = df[col].value_counts(dropna=False).reset_index()
        vc.columns = [col, "数量"]
        report_tables["司机分布"] = vc
        break

# 10. 起点终点流向分析（归一化后）
flow = df.groupby(["起点归一", "终点归一"]).size().reset_index(name="订单数").sort_values("订单数", ascending=False)
report_tables["流向统计"] = flow

# 11. 业务类型对比 (始终生成，如果无差异则为空表)
if "订单类型" in df.columns:
    diff_mask = (df["订单类型"] != df["业务类型_结构化"]) & (df["业务类型_结构化"] != "")
    df_diff = df.loc[diff_mask, ["订单概述", "业务类型_结构化", "业务类型_大类"]].rename(
        columns={"业务类型_结构化": "结构化业务类型", "业务类型_大类": "订单类型"}
    )
else:
    df_diff = pd.DataFrame(columns=["订单概述", "结构化业务类型", "订单类型"])
report_tables["业务类型对比"] = df_diff

# 12. 明细全表
report_tables["明细全表"] = df

# ===== 输出Excel，多sheet自动列宽和表格样式 =====
with pd.ExcelWriter(XLSX_OUT, engine="openpyxl", mode="w") as writer:
    for name, table in report_tables.items():
        table.to_excel(writer, index=False, sheet_name=name[:31])

def auto_adjust_column_width_and_style(xlsx_path: Path):
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2
        if ws.title != "明细全表":
            end_row = ws.max_row
            end_col = ws.max_column
            if end_row > 1 and end_col > 0:
                tab = Table(displayName=f"Table_{ws.title.replace(' ', '_')}",
                            ref=f"A1:{get_column_letter(end_col)}{end_row}")
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws.add_table(tab)
    wb.save(xlsx_path)

auto_adjust_column_width_and_style(XLSX_OUT)
logging.info("✅ 所有统计完成，分析Excel已输出到司机业务_统计分析.xlsx！")
