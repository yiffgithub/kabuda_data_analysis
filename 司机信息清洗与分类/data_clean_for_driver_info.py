# 请客服记录新数据时将其他(请说明地区）改为其他(请说明地区)
# 重构一下司机的编号

# 使用此程序前先手动在原有数据上重载司机编号，以及手动替换其他(请说明地区）改为其他(请说明地区)
r"""
clean_driver_data.py
~~~~~~~~~~~~~~~~~~~~
1) 读取「司机端—控制面板_司机档案.(csv|xlsx)」，完成字段清洗、标准化、派生标签，
   拆分多张主题表，并导出到 E:\kabuda_data_analysis\司机信息数据库。
2) 读取「司机面板模块化数据输出.xlsx」（只读，不覆盖），对第一张表：
    - 按 TARGET_IDS 清空/标记驾照
    - 标准化“活动地区”分隔符、去除“其他(请说明地区）”
   将处理结果作为一个新的 sheet 一并写入最终的 Excel。
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, Optional

import chardet
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ========== 配置 ==========
OUT_DIR = Path(r"E:\kabuda_data_analysis\司机信息数据库")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 主清洗输入文件
RAW_FILE = Path(r"E:\kabuda_data_analysis\Lark\司机端—控制面板_司机档案.csv")  # 或 .xlsx

# 最终输出
CSV_OUT = OUT_DIR / "司机信息.csv"
XLSX_OUT = OUT_DIR / "司机信息.xlsx"

# 面板模块化——只做读取，不写回
PANEL_INPUT = OUT_DIR / "司机面板模块化数据输出.xlsx"
TARGET_IDS = [
    'DR026', 'DR052', 'DR054', 'DR055', 'DR056', 'DR057', 'DR058',
    'DR059', 'DR062', 'DR063', 'DR064', 'DR150', 'DR219'
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)


# —— 1. 读取原始档案 —— #
def detect_encoding(fp: Path) -> str:
    with open(fp, "rb") as f:
        return chardet.detect(f.read())["encoding"] or "utf-8"


def load_raw(fp: Path) -> pd.DataFrame:
    ext = fp.suffix.lower()
    logging.info("加载原始档案: %s", fp)
    if ext == ".csv":
        enc = detect_encoding(fp)
        logging.info("检测到编码: %s", enc)
        return pd.read_csv(fp, encoding=enc, dtype=str)
    elif ext in (".xls", ".xlsx"):
        return pd.read_excel(fp, dtype=str, engine="openpyxl")
    else:
        raise ValueError(f"不支持的文件类型: {ext}")


# —— 2. 清洗管道 —— #
# 注意：这里已删除 “主活动地区”
COL_ORDER = [
    "自动编号", "司机等级", "名字", "提交时间", "提交人",
    "手机号", "邮箱", "微信号",
    "活动地区", "接单时间", "职业",
    "车型", "驾照",
    "活动地区-其他(请说明地区）-补充内容", "职业-其他-补充内容",
    "联系方式完整", "活跃度标签",
]


def normalize_na(df: pd.DataFrame) -> pd.DataFrame:
    for c in COL_ORDER:
        if c not in df:
            df[c] = np.nan
        df[c] = (df[c].astype(str)
                 .replace(["nan", "None", "NaN", ""], "缺失")
                 .str.strip()
                 .fillna("缺失"))
    return df


def clean_contact(df: pd.DataFrame) -> pd.DataFrame:
    phone_re = re.compile(r"^1\d{10}$|^6\d{9}$")
    # 提取纯数字手机号，校验格式
    df["手机号"] = (
        df["手机号"].astype(str)
        .apply(lambda x: re.sub(r"\D", "", x))
        .apply(lambda x: x if phone_re.match(x) else "缺失")
    )
    # 简单邮箱格式校验
    df["邮箱"] = df["邮箱"].apply(
        lambda x: x if re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", str(x)) else "缺失"
    )
    # 微信号长度校验
    df["微信号"] = df["微信号"].apply(
        lambda x: x if 3 < len(str(x)) < 50 and x != "缺失" else "缺失"
    )
    return df


def clean_dates(df: pd.DataFrame) -> pd.DataFrame:
    # 将“提交时间”统一为 YYYY-MM-DD，无法解析则标为“缺失”
    df["提交时间"] = (
        pd.to_datetime(df["提交时间"], errors="coerce")
        .dt.strftime("%Y-%m-%d")
        .fillna("缺失")
    )
    return df


# ========== 新增：清理“活动地区”中包含“其他(请说明地区)”的逻辑 ==========
def clean_area_other(df: pd.DataFrame) -> pd.DataFrame:
    """
    删除“活动地区”字段中出现的“其他(请说明地区)”子串：
    - 如果删除后变成空串，则填“缺失”；
    - 否则保留删除子串后剩余内容。
    """
    df["活动地区"] = (
        df["活动地区"]
        .astype(str)
        # 删除“其他(请说明地区)”字样
        .str.replace(r"其他\(请说明地区\)", "", regex=True)
        # 去掉前后空白
        .str.strip()
        # 如果空串，则标记“缺失”
        .replace({"": "缺失"})
    )
    return df


def clean_job_time(df: pd.DataFrame) -> pd.DataFrame:
    # 职业：不在列表内则标为“其它”或“缺失”
    valid_jobs = {"学生", "自由职业", "正在兼职", "全职", "待业"}
    df["职业"] = df["职业"].apply(
        # lambda x: x if x in valid_jobs else ("缺失" if x in ("缺失","nan","") else "其它")
        lambda x: x if x in valid_jobs else "缺失" if x in ("缺失", "nan", "") else "缺失"
    )
    # 接单时间：用第一个逗号前的时段，无法解析则“缺失”
    df["接单时间"] = (
        df["接单时间"].astype(str)
        .str.replace(r"[，、]", ",", regex=True)
        .apply(lambda x: "缺失" if x == "缺失" else x.split(",")[0])
    )
    return df


def clean_vehicle(df: pd.DataFrame) -> pd.DataFrame:
    # 车型 / 驾照：如果等于“缺失”或文件名后缀(.jpg/.png/.pdf)则标“缺失”，否则保留原值
    df["车型"] = df["车型"].astype(str).apply(
        lambda x: "缺失" if x.lower() == "缺失" or re.search(r"\.(jpg|png|pdf)$", x.lower()) else x
    )
    df["驾照"] = df["驾照"].astype(str).apply(
        lambda x: "缺失" if x.lower() == "" or x.lower() == " " else x
    )
    return df


def derive_tags(df: pd.DataFrame) -> pd.DataFrame:
    # “联系方式完整”：手机号/邮箱/微信号 都非“缺失”才标“完整”，否则“不完整”
    df["联系方式完整"] = df.apply(
        lambda r: "完整" if all(r[c] != "缺失" for c in ("手机号", "邮箱", "微信号")) else "不完整",
        axis=1,
    )
    # “活跃度标签”：如果“活动地区”非“缺失”且“接单时间”非“缺失”则“活跃”，否则“潜水”
    df["活跃度标签"] = df.apply(
        lambda r: "活跃" if r["活动地区"] != "缺失" and r["接单时间"] != "缺失" else "潜水",
        axis=1,
    )
    return df


def run_clean_pipeline(raw: pd.DataFrame) -> pd.DataFrame:
    return (
        raw.pipe(normalize_na)
        .pipe(clean_contact)
        .pipe(clean_dates)
        .pipe(clean_area_other)  # 先删“其他(请说明地区)”
        .pipe(clean_job_time)
        .pipe(clean_vehicle)
        .pipe(derive_tags)
    )[COL_ORDER]


# —— 3. 拆表 —— #
THEME_COLS: Dict[str, list[str]] = {
    "司机全量档案": COL_ORDER,
    "基础信息_司机": ["自动编号", "司机等级", "名字", "提交时间", "提交人"],
    "联系方式_司机": ["自动编号", "手机号", "邮箱", "微信号", "联系方式完整"],
    "业务信息_司机": ["自动编号", "活动地区", "接单时间", "职业", "活跃度标签"],
    "车辆信息_司机": ["自动编号", "车型", "驾照"],
    "补充信息_司机": ["自动编号", "活动地区-其他(请说明地区）-补充内容", "职业-其他-补充内容"],
}


def split_tables(df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    return {name: df[cols].drop_duplicates() for name, cols in THEME_COLS.items()}


# —— 4. 读取并处理“模块化面板” —— #
def load_and_process_panel(fp: Path) -> Optional[pd.DataFrame]:
    if not fp.exists():
        logging.info("模块化面板文件不存在，跳过 -> %s", fp)
        return None

    logging.info("加载模块化面板: %s", fp)
    all_s = pd.read_excel(fp, sheet_name=None, dtype=str)
    key = next(iter(all_s))
    df0 = all_s[key]

    # 清空/标记驾照
    mask = df0['自动编号'].isin(TARGET_IDS)
    df0.loc[mask, '驾照'] = ""
    df0.loc[~mask, '驾照'] = "已留存"

    # 标准化活动地区（删除“其他(请说明地区)”，空则“缺失”）
    df0['活动地区'] = (
        df0['活动地区'].astype(str)
        .str.replace(r'其他\(请说明地区\)', '', regex=True)
        .str.replace(r'[，、；;]', ',', regex=True)
        .str.replace(r',+', ',', regex=True)
        .str.strip(' ,')
        .replace({'': '缺失', 'nan': '缺失'})
    )
    return df0


# 将单元格做自适应列宽
def auto_adjust_column_width(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(xlsx_path)


# ========== 新增：为每个 sheet 套用 Excel Table 样式（可选） ==========
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def apply_excel_table_styles(xlsx_path: Path) -> None:
    """
    把每个 sheet 的数据区域转换为一个 Excel Table，并套用 TableStyleMedium9 样式。
    """
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        # 数据区域从 A1 到最右下角
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tbl = Table(displayName=f"{ws.title}_tbl", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
    wb.save(xlsx_path)


# ========== 新增：统计每个地区的司机数量 ==========
def count_by_region(df: pd.DataFrame) -> pd.DataFrame:
    """
    统计 DataFrame 中“活动地区”列下的司机数量。
    如果某个司机的“活动地区”包含多个地区（以英文逗号分隔），
    则该司机会被计入所有对应地区。
    返回一个 DataFrame，包含“活动地区”和对应的“司机数量”，按数量降序排序。
    """
    # 1. 确保“活动地区”列为字符串并按英文逗号拆分成列表
    temp = df.copy()
    temp["活动地区"] = temp["活动地区"].astype(str).str.split(",")

    # 2. 拆分后展开，每行只保留一个地区
    temp = temp.explode("活动地区")

    # 3. 去掉前后空白，如果拆分后是空串则标“缺失”
    temp["活动地区"] = (
        temp["活动地区"]
        .str.strip()
        .replace({"": "缺失"})
    )

    # 4. 按“活动地区”分组计数
    region_counts = (
        temp.groupby("活动地区")
        .size()
        .reset_index(name="司机数量")
        .sort_values("司机数量", ascending=False)
        .reset_index(drop=True)
    )

    return region_counts


# —— 5. 保存所有输出 —— #
def save_all(full: pd.DataFrame,
             tables: Dict[str, pd.DataFrame],
             panel_df: Optional[pd.DataFrame]) -> None:
    # 写出 CSV
    logging.info("写出 CSV -> %s", CSV_OUT)
    full.to_csv(CSV_OUT, index=False, encoding="utf-8-sig")

    # 写出 Excel 含多个 sheet
    logging.info("写出 Excel -> %s", XLSX_OUT)
    with pd.ExcelWriter(XLSX_OUT, engine="openpyxl", mode="w") as writer:
        # 主表
        full.to_excel(writer, sheet_name="司机全量档案", index=False)
        # 拆分的子表
        for nm, df_ in tables.items():
            if nm == "司机全量档案":
                continue
            df_.to_excel(writer, sheet_name=nm, index=False)
        # 模块化面板（如果有），新 sheet
        if panel_df is not None:
            panel_df.to_excel(writer, sheet_name="面板模块化输出", index=False)
        # 统计每个地区的司机数量
        # —— 新增：写入“地区统计”sheet —— #
        logging.info("写入地区统计 -> %s", XLSX_OUT)
        region_df = count_by_region(full)
        region_df.to_excel(writer, sheet_name="地区统计", index=False)
    # 自动调整列宽
    logging.info("调整列宽自适应内容宽度 -> %s", XLSX_OUT)
    auto_adjust_column_width(XLSX_OUT)

    # （可选）为所有 sheet 应用 Excel Table 样式
    logging.info("为所有 sheet 应用 Excel Table 样式 -> %s", XLSX_OUT)
    apply_excel_table_styles(XLSX_OUT)


def main():
    raw_df = load_raw(RAW_FILE)
    clean_df = run_clean_pipeline(raw_df)
    tables = split_tables(clean_df)

    panel_df = load_and_process_panel(PANEL_INPUT)
    save_all(clean_df, tables, panel_df)

    logging.info(
        "全部完成：%d 行，%d 名唯一司机。",
        len(clean_df),
        clean_df['自动编号'].nunique()
    )


if __name__ == "__main__":
    main()
